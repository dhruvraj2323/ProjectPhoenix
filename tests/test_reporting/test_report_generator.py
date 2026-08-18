from openpyxl import load_workbook

from reporting.report_generator import (
    ReportGenerator,
)

from reporting.reporting_models import (
    PerformanceSummary,
    TradeRecord,
)

def test_report_generator_m63_7_demo_observation_columns(
    tmp_path,
):
    """
    Verify M63.7 observation fields are written into
    the existing Trades sheet without changing the
    existing first 14 columns.
    """

    trade = TradeRecord(
        trade_id="M63-001",
        symbol="XAUUSDm",
        direction="BUY",
        strategy="S01_EMA_TREND",
        pattern="Bull Flag",
        entry_price=3350.0,
        exit_price=3360.0,
        stop_loss=3340.0,
        take_profit=3370.0,
        volume=0.10,
        profit_loss=100.0,
        status="CLOSED",

        strategy_decision="APPROVED",
        risk_decision="APPROVED",
        execution_status="EXECUTED",
        execution_message="Executed",
        execution_retcode=10009,

        requested_price=3349.8,
        executed_price=3350.0,

        requested_volume=0.10,
        executed_volume=0.10,

        order_check_retcode=0,
        order_check_message="Done",

        runtime_state="COMPLETED",
        trading_protection_state="ACTIVE",

        governance_state="APPROVED",
        governance_reason="NONE",

        balance=10000.0,
        equity=9950.0,
        free_margin=9000.0,

        open_positions=1,
        symbol_exposure=0.10,
        gross_exposure=0.10,
        net_exposure=0.10,
        portfolio_heat=1.0,

        risk_percent=1.0,
        drawdown=0.5,

        spread=0.20,
        slippage=0.20,

        mfe=12.0,
        mae=4.0,
    )

    summary = PerformanceSummary(
        total_trades=1,
        winning_trades=1,
        gross_profit=100.0,
        net_profit=100.0,
    )

    generator = ReportGenerator()

    original_directory = (
        ReportGenerator.REPORT_DIRECTORY
    )

    try:

        ReportGenerator.REPORT_DIRECTORY = (
            tmp_path
            / "reports"
            / "Daily"
        )

        report = generator.generate(
            [trade],
            summary,
        )

        workbook = load_workbook(
            report.output_file,
            read_only=True,
        )

        try:

            trades_sheet = workbook[
                "Trades"
            ]

            # Existing columns remain unchanged.
            assert trades_sheet["A1"].value == (
                "Trade ID"
            )

            assert trades_sheet["D1"].value == (
                "Strategy"
            )

            assert trades_sheet["K1"].value == (
                "Profit / Loss"
            )

            assert trades_sheet["D2"].value == (
                "S01_EMA_TREND"
            )

            # M63.7 fields.
            headers = {
                trades_sheet.cell(
                    row=1,
                    column=column,
                ).value: column
                for column in range(
                    1,
                    trades_sheet.max_column + 1,
                )
            }

            assert (
                headers[
                    "Strategy Decision"
                ]
                > 14
            )

            assert (
                headers[
                    "Risk Decision"
                ]
                > 14
            )

            assert (
                headers[
                    "Execution Status"
                ]
                > 14
            )

            assert (
                headers[
                    "Trading Protection State"
                ]
                > 14
            )

            assert (
                headers[
                    "Governance State"
                ]
                > 14
            )

            assert (
                headers[
                    "Risk Percent"
                ]
                > 14
            )

            assert (
                headers[
                    "Spread"
                ]
                > 14
            )

            assert (
                headers[
                    "Slippage"
                ]
                > 14
            )

            assert (
                headers[
                    "MFE"
                ]
                > 14
            )

            assert (
                headers[
                    "MAE"
                ]
                > 14
            )

            row = 2

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Strategy Decision"
                ],
            ).value == "APPROVED"

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Risk Decision"
                ],
            ).value == "APPROVED"

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Execution Status"
                ],
            ).value == "EXECUTED"

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Trading Protection State"
                ],
            ).value == "ACTIVE"

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Governance State"
                ],
            ).value == "APPROVED"

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Risk Percent"
                ],
            ).value == 1.0

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Spread"
                ],
            ).value == 0.20

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "Slippage"
                ],
            ).value == 0.20

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "MFE"
                ],
            ).value == 12.0

            assert trades_sheet.cell(
                row=row,
                column=headers[
                    "MAE"
                ],
            ).value == 4.0

        finally:
            workbook.close()

    finally:

        ReportGenerator.REPORT_DIRECTORY = (
            original_directory
        )