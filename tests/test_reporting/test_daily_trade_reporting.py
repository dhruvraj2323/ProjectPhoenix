"""
=================================================
Project Phoenix
Daily Trade Reporting Persistence Tests
Post-M63 Reporting Reliability
=================================================
"""
from datetime import UTC, datetime
from openpyxl import load_workbook
from deployment.execution_summary import (
    CycleExecutionSummary,
    SymbolExecutionResult,
    SymbolExecutionStatus,
)
from reporting.reporting_engine import (
    ReportingEngine,
)
from reporting.reporting_models import (
    TradeRecord,
)
def _trade(
    trade_id: str,
    symbol: str,
    profit_loss: float,
) -> TradeRecord:
    timestamp = datetime(
        2026,
        8,
        18,
        12,
        0,
        0,
        tzinfo=UTC,
    )
    return TradeRecord(
        trade_id=trade_id,
        symbol=symbol,
        direction="BUY",
        strategy="S01_EMA_TREND",
        pattern="",
        entry_price=100.0,
        exit_price=101.0,
        stop_loss=99.0,
        take_profit=102.0,
        volume=0.10,
        profit_loss=profit_loss,
        status="CLOSED",
        opened_at=timestamp,
        closed_at=timestamp,
    )
def _execution_summary(
    symbol: str,
    trade_id: str,
) -> CycleExecutionSummary:
    summary = CycleExecutionSummary()
    summary.add_result(
        SymbolExecutionResult(
            symbol=symbol,
            status=(
                SymbolExecutionStatus.EXECUTED
            ),
            trade_id=trade_id,
        )
    )
    return summary
def test_daily_reporting_accumulates_trades_across_cycles(
    tmp_path,
):
    engine = ReportingEngine()
    engine.generator.REPORT_DIRECTORY = (
        tmp_path
    )
    first_trade = _trade(
        trade_id="T001",
        symbol="EURUSDm",
        profit_loss=100.0,
    )
    second_trade = _trade(
        trade_id="T002",
        symbol="XAUUSDm",
        profit_loss=-25.0,
    )
    first_report = engine.run(
        trades=[first_trade],
        execution_summary=_execution_summary(
            symbol="EURUSDm",
            trade_id="T001",
        ),
    )
    assert len(
        first_report.trades
    ) == 1
    assert (
        first_report.summary.total_trades
        == 1
    )
    second_report = engine.run(
        trades=[second_trade],
        execution_summary=_execution_summary(
            symbol="XAUUSDm",
            trade_id="T002",
        ),
    )
    assert len(
        second_report.trades
    ) == 2
    assert (
        second_report.summary.total_trades
        == 2
    )
    assert (
        second_report.summary.net_profit
        == 75.0
    )
    workbook = load_workbook(
        second_report.output_file,
        read_only=True,
        data_only=True,
    )
    try:
        sheet = workbook[
            "Trades"
        ]
        assert (
            sheet.max_row
            == 3
        )
        assert (
            sheet["A2"].value
            == "T001"
        )
        assert (
            sheet["A3"].value
            == "T002"
        )
        assert (
            sheet["B2"].value
            == "EURUSDm"
        )
        assert (
            sheet["B3"].value
            == "XAUUSDm"
        )
    finally:
        workbook.close()
def test_daily_reporting_deduplicates_same_trade(
    tmp_path,
):
    engine = ReportingEngine()
    engine.generator.REPORT_DIRECTORY = (
        tmp_path
    )
    trade = _trade(
        trade_id="T100",
        symbol="BTCUSDm",
        profit_loss=50.0,
    )
    summary = _execution_summary(
        symbol="BTCUSDm",
        trade_id="T100",
    )
    first_report = engine.run(
        trades=[trade],
        execution_summary=summary,
    )
    second_report = engine.run(
        trades=[trade],
        execution_summary=summary,
    )
    assert len(
        first_report.trades
    ) == 1
    assert len(
        second_report.trades
    ) == 1
    assert (
        second_report.summary.total_trades
        == 1
    )
    assert (
        second_report.summary.net_profit
        == 50.0
    )
