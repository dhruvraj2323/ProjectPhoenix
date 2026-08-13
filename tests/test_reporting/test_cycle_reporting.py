"""
=================================================
Project Phoenix
Cycle Reporting Tests
M61.4 - Consolidated Cycle Reporting
=================================================
"""

from datetime import UTC, datetime

from openpyxl import load_workbook

from deployment.execution_summary import (
    CycleExecutionStatus,
    CycleExecutionSummary,
    SymbolExecutionResult,
    SymbolExecutionStatus,
)

from reporting.report_generator import (
    ReportGenerator,
)

from reporting.reporting_models import (
    PerformanceSummary,
    TradeRecord,
)


# =========================================================
# Helpers
# =========================================================

def _create_summary(
    results: list[
        SymbolExecutionResult
    ],
) -> CycleExecutionSummary:

    summary = CycleExecutionSummary()

    for result in results:
        summary.add_result(result)

    return summary


def _create_trade(
    symbol: str,
    trade_id: str,
) -> TradeRecord:

    execution_time = datetime(
        2026,
        8,
        13,
        12,
        0,
        0,
        tzinfo=UTC,
    )

    return TradeRecord(
        trade_id=trade_id,
        symbol=symbol,
        direction="BUY",
        strategy="S01_EMA_TREND",
        pattern="",
        entry_price=100.0,
        exit_price=0.0,
        stop_loss=90.0,
        take_profit=120.0,
        volume=0.01,
        profit_loss=0.0,
        status="OPEN",
        opened_at=execution_time,
        closed_at=execution_time,
    )


def _create_performance_summary() -> PerformanceSummary:

    return PerformanceSummary(
        total_trades=0,
        winning_trades=0,
        losing_trades=0,
        win_rate=0.0,
        gross_profit=0.0,
        gross_loss=0.0,
        net_profit=0.0,
        average_profit=0.0,
        average_loss=0.0,
        profit_factor=0.0,
    )


def _generate_report(
    tmp_path,
    trades: list[TradeRecord],
    execution_summary: CycleExecutionSummary,
):
    generator = ReportGenerator()

    generator.REPORT_DIRECTORY = (
        tmp_path
    )

    return generator.generate(
        trades=trades,
        summary=_create_performance_summary(),
        execution_summary=execution_summary,
    )


def _read_workbook(
    report,
):
    workbook = load_workbook(
        report.output_file,
        data_only=True,
    )

    return workbook


# =========================================================
# M61.4 Test A
# ALL_EXECUTED
# =========================================================

def test_cycle_reporting_all_executed(
    tmp_path,
):

    execution_summary = _create_summary(
        [
            SymbolExecutionResult(
                symbol="XAUUSDm",
                status=(
                    SymbolExecutionStatus.EXECUTED
                ),
                trade_id="XAU-001",
            ),
            SymbolExecutionResult(
                symbol="BTCUSDm",
                status=(
                    SymbolExecutionStatus.EXECUTED
                ),
                trade_id="BTC-001",
            ),
        ]
    )

    trades = [
        _create_trade(
            symbol="XAUUSDm",
            trade_id="XAU-001",
        ),
        _create_trade(
            symbol="BTCUSDm",
            trade_id="BTC-001",
        ),
    ]

    report = _generate_report(
        tmp_path=tmp_path,
        trades=trades,
        execution_summary=execution_summary,
    )

    workbook = _read_workbook(
        report,
    )

    try:

        assert {
            "Summary",
            "Trades",
            "Execution Summary",
        }.issubset(
            set(workbook.sheetnames)
        )

        sheet = workbook[
            "Execution Summary"
        ]

        assert (
            sheet["B4"].value
            == "ALL_EXECUTED"
        )

        assert (
            sheet["B5"].value
            == 2
        )

        assert (
            sheet["B6"].value
            == 2
        )

        assert (
            sheet["B7"].value
            == 0
        )

        assert (
            sheet["B8"].value
            == 0
        )

        assert (
            sheet["A11"].value
            == "XAUUSDm"
        )

        assert (
            sheet["B11"].value
            == "EXECUTED"
        )

        assert (
            sheet["C11"].value
            == "XAU-001"
        )

        assert (
            sheet["A12"].value
            == "BTCUSDm"
        )

        assert (
            sheet["B12"].value
            == "EXECUTED"
        )

        assert (
            sheet["C12"].value
            == "BTC-001"
        )

    finally:

        workbook.close()


# =========================================================
# M61.4 Test B
# PARTIAL_SUCCESS
# =========================================================

def test_cycle_reporting_partial_success(
    tmp_path,
):

    execution_summary = _create_summary(
        [
            SymbolExecutionResult(
                symbol="XAUUSDm",
                status=(
                    SymbolExecutionStatus.FAILED
                ),
                error=(
                    "Simulated XAUUSDm failure."
                ),
            ),
            SymbolExecutionResult(
                symbol="BTCUSDm",
                status=(
                    SymbolExecutionStatus.EXECUTED
                ),
                trade_id="BTC-002",
            ),
        ]
    )

    trades = [
        _create_trade(
            symbol="BTCUSDm",
            trade_id="BTC-002",
        ),
    ]

    report = _generate_report(
        tmp_path=tmp_path,
        trades=trades,
        execution_summary=execution_summary,
    )

    workbook = _read_workbook(
        report,
    )

    try:

        sheet = workbook[
            "Execution Summary"
        ]

        assert (
            sheet["B4"].value
            == "PARTIAL_SUCCESS"
        )

        assert (
            sheet["B5"].value
            == 2
        )

        assert (
            sheet["B6"].value
            == 1
        )

        assert (
            sheet["B7"].value
            == 0
        )

        assert (
            sheet["B8"].value
            == 1
        )

        assert (
            sheet["A11"].value
            == "XAUUSDm"
        )

        assert (
            sheet["B11"].value
            == "FAILED"
        )

        assert (
            "Simulated XAUUSDm failure."
            in sheet["D11"].value
        )

        assert (
            sheet["A12"].value
            == "BTCUSDm"
        )

        assert (
            sheet["B12"].value
            == "EXECUTED"
        )

        assert (
            sheet["C12"].value
            == "BTC-002"
        )

    finally:

        workbook.close()


# =========================================================
# M61.4 Test C
# NO_TRADES
# =========================================================

def test_cycle_reporting_no_trades(
    tmp_path,
):

    execution_summary = _create_summary(
        [
            SymbolExecutionResult(
                symbol="XAUUSDm",
                status=(
                    SymbolExecutionStatus.NO_TRADE
                ),
            ),
            SymbolExecutionResult(
                symbol="BTCUSDm",
                status=(
                    SymbolExecutionStatus.NO_TRADE
                ),
            ),
        ]
    )

    report = _generate_report(
        tmp_path=tmp_path,
        trades=[],
        execution_summary=execution_summary,
    )

    workbook = _read_workbook(
        report,
    )

    try:

        sheet = workbook[
            "Execution Summary"
        ]

        assert (
            sheet["B4"].value
            == "NO_TRADES"
        )

        assert (
            sheet["B5"].value
            == 2
        )

        assert (
            sheet["B6"].value
            == 0
        )

        assert (
            sheet["B7"].value
            == 2
        )

        assert (
            sheet["B8"].value
            == 0
        )

        assert (
            sheet["A11"].value
            == "XAUUSDm"
        )

        assert (
            sheet["B11"].value
            == "NO_TRADE"
        )

        assert (
            sheet["A12"].value
            == "BTCUSDm"
        )

        assert (
            sheet["B12"].value
            == "NO_TRADE"
        )

    finally:

        workbook.close()


# =========================================================
# M61.4 Test D
# ALL_FAILED
# =========================================================

def test_cycle_reporting_all_failed(
    tmp_path,
):

    execution_summary = _create_summary(
        [
            SymbolExecutionResult(
                symbol="XAUUSDm",
                status=(
                    SymbolExecutionStatus.FAILED
                ),
                error=(
                    "XAUUSDm execution failed."
                ),
            ),
            SymbolExecutionResult(
                symbol="BTCUSDm",
                status=(
                    SymbolExecutionStatus.FAILED
                ),
                error=(
                    "BTCUSDm execution failed."
                ),
            ),
        ]
    )

    report = _generate_report(
        tmp_path=tmp_path,
        trades=[],
        execution_summary=execution_summary,
    )

    workbook = _read_workbook(
        report,
    )

    try:

        sheet = workbook[
            "Execution Summary"
        ]

        assert (
            sheet["B4"].value
            == "ALL_FAILED"
        )

        assert (
            sheet["B5"].value
            == 2
        )

        assert (
            sheet["B6"].value
            == 0
        )

        assert (
            sheet["B7"].value
            == 0
        )

        assert (
            sheet["B8"].value
            == 2
        )

        assert (
            sheet["A11"].value
            == "XAUUSDm"
        )

        assert (
            sheet["B11"].value
            == "FAILED"
        )

        assert (
            "XAUUSDm execution failed."
            in sheet["D11"].value
        )

        assert (
            sheet["A12"].value
            == "BTCUSDm"
        )

        assert (
            sheet["B12"].value
            == "FAILED"
        )

        assert (
            "BTCUSDm execution failed."
            in sheet["D12"].value
        )

    finally:

        workbook.close()


# =========================================================
# M61.4 Test E
# Empty Summary Compatibility
# =========================================================

def test_cycle_reporting_without_execution_summary(
    tmp_path,
):

    generator = ReportGenerator()

    generator.REPORT_DIRECTORY = (
        tmp_path
    )

    report = generator.generate(
        trades=[],
        summary=_create_performance_summary(),
    )

    workbook = _read_workbook(
        report,
    )

    try:

        assert {
            "Summary",
            "Trades",
            "Execution Summary",
        }.issubset(
            set(workbook.sheetnames)
        )

        sheet = workbook[
            "Execution Summary"
        ]

        assert (
            sheet["B4"].value
            == "Not Available"
        )

    finally:

        workbook.close()