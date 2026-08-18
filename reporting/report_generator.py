"""
=================================================
Project Phoenix
Report Generator
M63.7 - Demo Reporting & Analytics
=================================================
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from reporting.reporting_models import (
    DailyReport,
    PerformanceSummary,
    TradeRecord,
)


class ReportGenerator:
    """
    Builds and writes Daily Trading Reports.

    Existing responsibilities:
    - Build DailyReport model
    - Create XLSX workbook
    - Write performance summary
    - Write trade records
    - Write cycle execution summary

    M63.7:
    - Preserve existing reporting architecture
    - Extend Trades sheet with DEMO observation data
    """

    REPORT_DIRECTORY = Path(
        "reports"
    ) / "Daily"

    # --------------------------------------------------
    # Generate
    # --------------------------------------------------

    def generate(
        self,
        trades: list[TradeRecord],
        summary: PerformanceSummary,
        execution_summary: Any = None,
    ) -> DailyReport:
        """
        Generate complete daily report.
        """

        report = DailyReport()

        report.trades = trades

        report.summary = summary

        report.execution_summary = (
            execution_summary
        )

        report.report_date = datetime.now(
            UTC
        )

        report.generated_at = datetime.now(
            UTC
        )

        report.report_name = (
            report.report_date.strftime(
                "%Y-%m-%d"
            )
            + "_Trading_Report"
        )

        self.REPORT_DIRECTORY.mkdir(
            parents=True,
            exist_ok=True,
        )

        output_path = (
            self.REPORT_DIRECTORY
            / f"{report.report_name}.xlsx"
        )

        report.output_file = str(
            output_path
        )

        self._write_workbook(
            trades=trades,
            summary=summary,
            execution_summary=execution_summary,
            output_path=output_path,
            report_date=report.report_date,
            generated_at=report.generated_at,
        )

        return report

    # --------------------------------------------------
    # Workbook
    # --------------------------------------------------

    def _write_workbook(
        self,
        trades: list[TradeRecord],
        summary: PerformanceSummary,
        execution_summary: Any,
        output_path: Path,
        report_date: datetime,
        generated_at: datetime,
    ) -> None:

        workbook = Workbook()

        # --------------------------------------------------
        # Sheets
        # --------------------------------------------------

        summary_sheet = workbook.active

        summary_sheet.title = "Summary"

        trades_sheet = workbook.create_sheet(
            title="Trades",
        )

        execution_sheet = (
            workbook.create_sheet(
                title="Execution Summary",
            )
        )

        # --------------------------------------------------
        # Write
        # --------------------------------------------------

        self._write_summary_sheet(
            worksheet=summary_sheet,
            summary=summary,
            report_date=report_date,
            generated_at=generated_at,
        )

        self._write_trades_sheet(
            worksheet=trades_sheet,
            trades=trades,
        )

        self._write_execution_summary_sheet(
            worksheet=execution_sheet,
            execution_summary=execution_summary,
        )

        # --------------------------------------------------
        # Auto Size
        # --------------------------------------------------

        self._auto_size_columns(
            summary_sheet,
        )

        self._auto_size_columns(
            trades_sheet,
        )

        self._auto_size_columns(
            execution_sheet,
        )

        # --------------------------------------------------
        # Save
        # --------------------------------------------------

        workbook.save(
            output_path,
        )

        workbook.close()

    # --------------------------------------------------
    # Summary Sheet
    # --------------------------------------------------

    def _write_summary_sheet(
        self,
        worksheet,
        summary: PerformanceSummary,
        report_date: datetime,
        generated_at: datetime,
    ) -> None:

        worksheet["A1"] = (
            "Project Phoenix"
        )

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet["A2"] = (
            "Daily Trading Report"
        )

        worksheet["A2"].font = Font(
            bold=True,
            size=13,
        )

        worksheet["A4"] = (
            "Report Date"
        )

        worksheet["B4"] = (
            report_date.strftime(
                "%Y-%m-%d"
            )
        )

        worksheet["A5"] = (
            "Generated At"
        )

        worksheet["B5"] = (
            generated_at.strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )

        worksheet["A7"] = (
            "Metric"
        )

        worksheet["B7"] = (
            "Value"
        )

        worksheet["A7"].font = Font(
            bold=True,
        )

        worksheet["B7"].font = Font(
            bold=True,
        )

        metrics = [
            (
                "Total Trades",
                summary.total_trades,
            ),
            (
                "Winning Trades",
                summary.winning_trades,
            ),
            (
                "Losing Trades",
                summary.losing_trades,
            ),
            (
                "Win Rate (%)",
                summary.win_rate,
            ),
            (
                "Gross Profit",
                summary.gross_profit,
            ),
            (
                "Gross Loss",
                summary.gross_loss,
            ),
            (
                "Net Profit",
                summary.net_profit,
            ),
            (
                "Average Profit",
                summary.average_profit,
            ),
            (
                "Average Loss",
                summary.average_loss,
            ),
            (
                "Profit Factor",
                summary.profit_factor,
            ),
        ]

        start_row = 8

        for index, (
            metric,
            value,
        ) in enumerate(
            metrics,
            start=start_row,
        ):

            worksheet.cell(
                row=index,
                column=1,
                value=metric,
            )

            worksheet.cell(
                row=index,
                column=2,
                value=value,
            )

    # --------------------------------------------------
    # Trades Sheet
    # --------------------------------------------------

    def _write_trades_sheet(
        self,
        worksheet,
        trades: list[TradeRecord],
    ) -> None:
        """
        Write existing trade fields followed by
        M63.7 DEMO observation fields.

        Existing first 14 columns remain unchanged
        for backward compatibility.
        """

        headers = [

            # ------------------------------------------
            # Existing M57 fields
            # ------------------------------------------

            "Trade ID",
            "Symbol",
            "Direction",
            "Strategy",
            "Pattern",
            "Entry Price",
            "Exit Price",
            "Stop Loss",
            "Take Profit",
            "Volume",
            "Profit / Loss",
            "Status",
            "Opened At",
            "Closed At",

            # ------------------------------------------
            # M63.7 Demo Observation
            # ------------------------------------------

            "Strategy Decision",
            "Risk Decision",
            "Execution Status",
            "Execution Message",
            "Execution Retcode",
            "Requested Price",
            "Executed Price",
            "Requested Volume",
            "Executed Volume",
            "Order Check Retcode",
            "Order Check Message",

            "Runtime State",
            "Trading Protection State",

            "Governance State",
            "Governance Reason",
            "Balance",
            "Equity",
            "Free Margin",
            "Open Positions",
            "Symbol Exposure",
            "Gross Exposure",
            "Net Exposure",
            "Portfolio Heat",
            "Risk Percent",
            "Drawdown",

            "Spread",
            "Slippage",
            "MFE",
            "MAE",

            "Observation Error",
        ]

        # --------------------------------------------------
        # Header
        # --------------------------------------------------

        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True,
            )

        # --------------------------------------------------
        # Rows
        # --------------------------------------------------

        for row_index, trade in enumerate(
            trades,
            start=2,
        ):

            values = [

                # ------------------------------------------
                # Existing fields
                # ------------------------------------------

                trade.trade_id,
                trade.symbol,
                trade.direction,
                trade.strategy,
                trade.pattern,
                trade.entry_price,
                trade.exit_price,
                trade.stop_loss,
                trade.take_profit,
                trade.volume,
                trade.profit_loss,
                trade.status,
                self._format_datetime(
                    trade.opened_at
                ),
                self._format_datetime(
                    trade.closed_at
                ),

                # ------------------------------------------
                # M63.7 observation
                # ------------------------------------------

                trade.strategy_decision,
                trade.risk_decision,
                trade.execution_status,
                trade.execution_message,
                trade.execution_retcode,
                trade.requested_price,
                trade.executed_price,
                trade.requested_volume,
                trade.executed_volume,
                trade.order_check_retcode,
                trade.order_check_message,

                trade.runtime_state,
                trade.trading_protection_state,

                trade.governance_state,
                trade.governance_reason,
                trade.balance,
                trade.equity,
                trade.free_margin,
                trade.open_positions,
                trade.symbol_exposure,
                trade.gross_exposure,
                trade.net_exposure,
                trade.portfolio_heat,
                trade.risk_percent,
                trade.drawdown,

                trade.spread,
                trade.slippage,
                trade.mfe,
                trade.mae,

                trade.observation_error,
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

    # --------------------------------------------------
    # Execution Summary
    # --------------------------------------------------

    def _write_execution_summary_sheet(
        self,
        worksheet,
        execution_summary: Any,
    ) -> None:
        """
        Write cycle-level execution summary.

        Reporting remains independent from the
        deployment package by accepting Any.
        """

        worksheet["A1"] = (
            "Project Phoenix"
        )

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet["A2"] = (
            "Cycle Execution Summary"
        )

        worksheet["A2"].font = Font(
            bold=True,
            size=13,
        )

        if execution_summary is None:

            worksheet["A4"] = (
                "Execution Summary"
            )

            worksheet["B4"] = (
                "Not Available"
            )

            return

        worksheet["A4"] = (
            "Cycle Status"
        )

        worksheet["B4"] = (
            self._enum_value(
                execution_summary.status
            )
        )

        worksheet["A5"] = (
            "Total Symbols"
        )

        worksheet["B5"] = (
            execution_summary.total_symbols
        )

        worksheet["A6"] = (
            "Executed Symbols"
        )

        worksheet["B6"] = (
            execution_summary.executed_symbols
        )

        worksheet["A7"] = (
            "No Trade Symbols"
        )

        worksheet["B7"] = (
            execution_summary.no_trade_symbols
        )

        worksheet["A8"] = (
            "Failed Symbols"
        )

        worksheet["B8"] = (
            execution_summary.failed_symbols
        )

        worksheet["A10"] = (
            "Symbol"
        )

        worksheet["B10"] = (
            "Status"
        )

        worksheet["C10"] = (
            "Trade ID"
        )

        worksheet["D10"] = (
            "Error"
        )

        for column in (
            "A10",
            "B10",
            "C10",
            "D10",
        ):

            worksheet[column].font = Font(
                bold=True,
            )

        results = getattr(
            execution_summary,
            "symbol_results",
            [],
        )

        for row_index, result in enumerate(
            results,
            start=11,
        ):

            worksheet.cell(
                row=row_index,
                column=1,
                value=getattr(
                    result,
                    "symbol",
                    "",
                ),
            )

            worksheet.cell(
                row=row_index,
                column=2,
                value=self._enum_value(
                    getattr(
                        result,
                        "status",
                        "",
                    )
                ),
            )

            worksheet.cell(
                row=row_index,
                column=3,
                value=getattr(
                    result,
                    "trade_id",
                    "",
                ),
            )

            worksheet.cell(
                row=row_index,
                column=4,
                value=getattr(
                    result,
                    "error",
                    "",
                ),
            )

    # --------------------------------------------------
    # Utilities
    # --------------------------------------------------

    @staticmethod
    def _enum_value(
        value: Any,
    ) -> str:

        if value is None:
            return ""

        enum_value = getattr(
            value,
            "value",
            None,
        )

        if enum_value is not None:
            return str(enum_value)

        return str(value)

    @staticmethod
    def _format_datetime(
        value: Any,
    ) -> str:

        if value is None:
            return ""

        if hasattr(
            value,
            "strftime",
        ):

            return value.strftime(
                "%Y-%m-%d %H:%M:%S"
            )

        return str(value)

    def _auto_size_columns(
        self,
        worksheet,
    ) -> None:

        for column_cells in (
            worksheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column,
                )
            )

            for cell in column_cells:

                value = cell.value

                if value is None:
                    continue

                value_length = len(
                    str(value)
                )

                if value_length > max_length:
                    max_length = (
                        value_length
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40,
            )